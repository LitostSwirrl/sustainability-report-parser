"""
XlsxManager — read/write local xlsx file for the active agentic workflow.

Replaces Google Sheets (gspread) dependency for the 5 active scripts:
  - list_companies.py (READ)
  - check_progress.py (READ)
  - publish_to_xlsx.py (WRITE)
  - write_group_results.py (WRITE)
  - merge_groups.py (WRITE)
"""

import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from .config import XLSX_PATH, XLSX_COMPANY_LIST_TAB
from .utils import get_logger


class XlsxManager:
    """Manage reads/writes to the local xlsx record spreadsheet."""

    def __init__(self, xlsx_path: Optional[Path] = None) -> None:
        self.xlsx_path = xlsx_path or XLSX_PATH
        self.logger = get_logger()

    # ------------------------------------------------------------------
    # READ helpers
    # ------------------------------------------------------------------

    def _load_workbook_readonly(self) -> Workbook:
        """Open the workbook in read-only + data-only mode for performance."""
        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"xlsx file not found: {self.xlsx_path}")
        return load_workbook(self.xlsx_path, read_only=True, data_only=True)

    def get_tab_names(self) -> list[str]:
        """Return all tab (worksheet) names in the workbook."""
        wb = self._load_workbook_readonly()
        names = wb.sheetnames
        wb.close()
        return names

    def read_sheet_as_dicts(self, tab_name: str) -> list[dict]:
        """
        Read a tab's rows as a list of dicts keyed by the header row.

        Uses read_only=True, data_only=True for performance.
        """
        wb = self._load_workbook_readonly()
        if tab_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Tab '{tab_name}' not found in {self.xlsx_path}")

        ws = wb[tab_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        headers = [str(h) if h is not None else "" for h in rows[0]]
        result: list[dict] = []
        for row in rows[1:]:
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue
            record = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                val = row[i] if i < len(row) else None
                record[header] = val if val is not None else ""
            result.append(record)

        return result

    def get_company_list(
        self,
        filter_to_analyze: bool = True,
        year: Optional[str] = None,
        industry: Optional[str] = None,
        tab_name: Optional[str] = None,
    ) -> list[dict]:
        """
        Load company list from the xlsx — drop-in replacement for Sheets loading.

        Args:
            filter_to_analyze: Only return rows where 待分析 is truthy.
            year: Filter by 年度 value.
            industry: Filter by 產業別 value.
            tab_name: Override the default company list tab name.

        Returns:
            List of company dicts with normalized string keys.
        """
        target_tab = tab_name or XLSX_COMPANY_LIST_TAB
        data = self.read_sheet_as_dicts(target_tab)
        self.logger.info(f"Loaded {len(data)} rows from xlsx tab '{target_tab}'")

        companies: list[dict] = []
        for row in data:
            if filter_to_analyze:
                flag = str(row.get("待分析", "")).strip().upper()
                if flag not in ("TRUE", "1", "YES", "Y"):
                    continue

            year_val = str(row.get("年度", "")).strip()
            industry_val = str(row.get("產業別", "")).strip()

            if year and year_val != year:
                continue
            if industry and industry_val != industry:
                continue

            companies.append(
                {
                    "company_code": str(row.get("公司代碼", "")).strip(),
                    "company_name": str(row.get("公司簡稱", "")).strip(),
                    "full_name": str(row.get("公司全名", "")).strip(),
                    "industry": industry_val,
                    "market": str(row.get("市場別", "")).strip(),
                    "year": year_val,
                    "file_link": str(row.get("檔案連結", "")).strip(),
                    "file_size_mb": str(row.get("檔案大小(MB)", "")).strip(),
                    "download_status": str(row.get("下載狀態", "")).strip(),
                    "to_analyze": str(row.get("待分析", "")).strip(),
                    "notes": str(row.get("備註", "")).strip(),
                    "last_updated": str(row.get("最後更新時間", "")).strip(),
                }
            )

        return companies

    # ------------------------------------------------------------------
    # WRITE helpers
    # ------------------------------------------------------------------

    def _load_workbook_write(self) -> Workbook:
        """Open the workbook in full write mode, creating a .bak backup first."""
        if not self.xlsx_path.exists():
            raise FileNotFoundError(f"xlsx file not found: {self.xlsx_path}")

        # Create timestamped backup
        bak_path = self.xlsx_path.with_suffix(
            f".{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak"
        )
        shutil.copy2(self.xlsx_path, bak_path)
        self.logger.info(f"Backup created: {bak_path}")

        return load_workbook(self.xlsx_path)

    def _ensure_tab_exists(self, wb: Workbook, tab_name: str) -> Worksheet:
        """Get an existing worksheet or create a new one."""
        if tab_name in wb.sheetnames:
            return wb[tab_name]
        ws = wb.create_sheet(title=tab_name)
        self.logger.info(f"Created new tab: '{tab_name}'")
        return ws

    def _format_header_row(self, ws: Worksheet, num_cols: int) -> None:
        """Bold the first row and freeze panes below it."""
        bold_font = Font(bold=True)
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold_font
        ws.freeze_panes = "A2"

    def clear_and_write_tab(
        self,
        tab_name: str,
        headers: list[str],
        rows: list[list[str]],
    ) -> int:
        """
        Delete tab contents, write headers + rows, apply formatting.

        Args:
            tab_name: Target tab name (created if missing).
            headers: Header row values.
            rows: Data rows.

        Returns:
            Number of data rows written.
        """
        wb = self._load_workbook_write()

        # Remove old tab if it exists, then recreate
        if tab_name in wb.sheetnames:
            del wb[tab_name]
            self.logger.info(f"Deleted existing tab: '{tab_name}'")
        ws = wb.create_sheet(title=tab_name)

        # Write header
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        self._format_header_row(ws, len(headers))

        wb.save(self.xlsx_path)
        wb.close()
        self.logger.info(
            f"clear_and_write_tab: wrote {len(rows)} rows to '{tab_name}'"
        )
        return len(rows)

    def append_rows_to_tab(
        self,
        tab_name: str,
        headers: list[str],
        rows: list[list[str]],
    ) -> int:
        """
        Append rows to an existing tab (create with headers if missing).

        Args:
            tab_name: Target tab name.
            headers: Header row (used only if tab is created fresh).
            rows: Data rows to append.

        Returns:
            Number of rows appended.
        """
        if not rows:
            self.logger.warning("append_rows_to_tab: no rows to append")
            return 0

        wb = self._load_workbook_write()
        ws = self._ensure_tab_exists(wb, tab_name)

        # If the tab is brand-new (empty), write headers first
        if ws.max_row is None or ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            self._format_header_row(ws, len(headers))
            start_row = 2
        else:
            start_row = ws.max_row + 1

        for row_offset, row in enumerate(rows):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=start_row + row_offset, column=col_idx, value=value)

        wb.save(self.xlsx_path)
        wb.close()
        self.logger.info(
            f"append_rows_to_tab: appended {len(rows)} rows to '{tab_name}' "
            f"(starting at row {start_row})"
        )
        return len(rows)

    def write_explanation_tab(
        self,
        tab_name: str,
        rows: list[list[str]],
    ) -> int:
        """
        Write explanation tab with special formatting.

        Formatting: bold title (row 1, font size 14), bold section headers
        (Chinese numbered headings), wide column A (100 chars), text wrap.

        Args:
            tab_name: Target tab name.
            rows: List of single-element lists (one column).

        Returns:
            Number of rows written.
        """
        wb = self._load_workbook_write()

        if tab_name in wb.sheetnames:
            del wb[tab_name]
        ws = wb.create_sheet(title=tab_name)

        # Write rows
        for row_idx, row in enumerate(rows, start=1):
            value = row[0] if row else ""
            ws.cell(row=row_idx, column=1, value=value)

        # Format title row
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)

        # Format section headers (一、二、三…)
        cn_number_chars = "一二三四五六七八九十"
        for row_idx, row in enumerate(rows, start=1):
            text = row[0] if row else ""
            if text and len(text) > 1 and text[0] in cn_number_chars:
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)

        # Set column A width (approximately 100 characters)
        ws.column_dimensions["A"].width = 100

        wb.save(self.xlsx_path)
        wb.close()
        self.logger.info(
            f"write_explanation_tab: wrote {len(rows)} rows to '{tab_name}'"
        )
        return len(rows)
